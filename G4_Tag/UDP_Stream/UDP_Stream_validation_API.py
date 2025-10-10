import socket
import struct
import datetime
import time
import serial
import logging
import json
import re
from openpyxl import load_workbook
#from Report_process import RateProcessor
#from Profile import JSONMessageApp
#from Tcp_command_1 import DevCommandSender
from Power_supply import KoradPowerSupply
g4_tag1_id = 8457892
server_ip = "192.168.1.5"
PS = KoradPowerSupply(config_file="config.json")
PS.connect()
PS.set_voltage(0)
PS.disconnect()
time.sleep(5)
PS.connect()
PS.set_voltage(3)
#TCP = JSONMessageApp(server_ip)
#Report_processor = RateProcessor()
#sender = DevCommandSender()

def load_config():
    with open("config.json", "r") as file:
        return json.load(file)

# Read Arduino serial port from config
config = load_config()
Arduino_serial_port = config["Arduino_serial_port"]
host = config["host"]
port = config["port"]
timeout = config["timeout"]


def Report(test_id, f):
    file = 'G4_tag_testcases_automation.xlsx'
    STP = load_workbook(file)
    sheet = STP['RF_900MHz']

    for row in range(1, sheet.max_row + 1):
        test_case = sheet[f'D{row}'].value
        if test_case and test_id in str(test_case):
            result_cell = f'F{row}'
            sheet[result_cell].value = "PASS" if f == 1 else "FAIL"
            print(f"Updated row {row} -> {sheet[result_cell].value}")

    STP.save(file)


# Configure logging
LOG_FILE = "output.log"  # Predefined file name

def Output(data: str):
    """Writes data to a predefined file with a timestamp, each entry on a new line."""
    time_stamp = datetime.datetime.now()
    with open(LOG_FILE, "a") as file:  # Append mode to keep old data
        file.write(f"{time_stamp} - {data}\n")  # Write timestamp + data + newline

logging.basicConfig(
    filename='Debug_log.log',  # Log file name
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def get_tester_feedback(message):
    while True:
        print(message)
        feedback = input("Please type 'yes' or 'no': ").strip().lower()
        if feedback == 'yes':
            return True
        elif feedback == 'no':
            return False
        else:
            print("Invalid input. Try again.\n")

class SerialConnection:
    def __init__(self, port, baud_rate, timeout=1):
        self.port = port
        self.baud_rate = baud_rate
        self.timeout = timeout
        self.ser = None

    def init_serial(self):
        try:
            self.ser = serial.Serial(self.port, self.baud_rate, timeout=self.timeout)
            if self.ser.isOpen():
                print(f"Connected to {self.port}")
            # Wait for Arduino to initialize
            time.sleep(2)
        except serial.SerialException as e:
            print(f"Error opening serial port: {e}")
    #Supriya
    def readline(self):
        if self.ser is None:
            # This will give you a clear, early error if serial isn't initialized
            raise RuntimeError("Serial connection not initialized. Call init_serial() first.")

        if self.ser.in_waiting:
            return self.ser.readline().decode(errors="ignore").strip()

        return None

    '''def in_waiting(self):
        return self.ser.in_waiting

    def readline(self):
        if self.ser.in_waiting:
            return self.ser.readline().decode().strip()
        return None
    '''
    #supriya end
    def send_command(self, command):
        try:
            if self.ser:
                # Send the command
                self.ser.write(f"{command}\n".encode())  # Encode the string to bytes and send with newline
                logging.info(command)
                print(f"Sent: {command}")
                # Wait for response
                time.sleep(1)  # Wait for response
                if self.ser.in_waiting > 0:
                    response = self.ser.read(self.ser.in_waiting).decode('utf-8').strip()
                    print(f"Arduino: {response}")
        except serial.SerialException as e:
            print(f"Error communicating with Arduino: {e}")

    def close(self):
        if self.ser and self.ser.isOpen():
            self.ser.close()
            print("Serial connection closed.")



class PacketDecoder:
    @staticmethod
    def calculate_checksum(data):
        """
        Calculates the checksum by summing all bytes in the data.
        """
        return sum(data) & 0xFFFF  # Ensure it is a 2-byte value
    @staticmethod
    def decode_header(header):
        if len(header) < 13:
            print(f"Error: Header too short (length {len(header)})")
            return None  # Return None to indicate an invalid packet

        cycle_counter = header[0]
        star_mac_id = ':'.join(f'{b:02X}' for b in header[1:7])
        data_length = struct.unpack('<H', header[7:9])[0]  # Little-endian
        data_checksum = struct.unpack('<H', header[9:11])[0]  # Little-endian
        header_checksum = struct.unpack('<H', header[11:13])[0]  # Little-endian

        print(f"Cycle Counter: {cycle_counter}")
        print(f"Star MAC Id: {star_mac_id}")
        print(f"Data Length: {data_length}")
        print(f"Data Checksum: {data_checksum}")
        print(f"Header Checksum: {header_checksum}")

        return data_length,data_checksum

    @staticmethod
    def decode_status_byte(status_byte):
        if not isinstance(status_byte, int) or status_byte > 255:
            print(f"Error: Invalid status byte {status_byte}")
            return None  # Return None if the byte is invalid
        bit_fields = f'{status_byte:08b}'
        print(f"Status Byte: {bit_fields}")
        print(f"  Button 4: {bit_fields[0]}")
        print(f"  Button 3: {bit_fields[1]}")
        print(f"  Button 2: {bit_fields[2]}")
        print(f"  Button 1: {bit_fields[3]}")
        print(f"  Motion Flag: {bit_fields[4]}")
        print(f"  Retry Count: {int(bit_fields[5:7], 2)}")  # Convert bits 5 and 6 to decimal
        print(f"  Reserved: {bit_fields[7]}")
        return {
            'Status Byte': bit_fields,
            'Button 4': bit_fields[0],
            'Button 3': bit_fields[1],
            'Button 2': bit_fields[2],
            'Button 1': bit_fields[3],
            'Motion Flag': bit_fields[4],
            'Retry Count': int(bit_fields[5:7], 2),
            'Reserved': bit_fields[7]
        }

    @staticmethod
    def decode_Alive_status(Alive_status):
        if not isinstance(Alive_status, int) or Alive_status > 255:
            print(f"Error: Invalid status byte {Alive_status}")
            return None  # Return None if the byte is invalid
        bit_fields = f'{Alive_status:08b}'
        print(f"  Sleep status: {bit_fields[3]}")
        print(f"  Retry Count: {int(bit_fields[4:8], 2)}")  # Convert bits 5 and 8 to decimal

        return {
            'Sleep_status' : bit_fields[3],
            'Index': int(bit_fields[4:8], 2)
        }

    @staticmethod
    def decode_location_packet(packet,expected_checksum):
        if len(packet) < 34:
            print(f"Error: Location packet too short! Length: {len(packet)}")
            return None  # Return None if the packet is too short

        computed_checksum = PacketDecoder.calculate_checksum(packet)
        if computed_checksum != expected_checksum:
            print(
                f"Error: Data Checksum Mismatch! Computed: {computed_checksum:04X}, Expected: {expected_checksum:04X}")

        else:
            print(f"Data Checksum Matched: {computed_checksum:04X}")

        device_type = packet[0]
        tag_id_raw = int.from_bytes(packet[1:5], byteorder='little')
        tag_id = tag_id_raw & 0x0FFFFFFF  # Remove the MSB nibble
        raw_rssi = packet[8]
        rssi = (raw_rssi - 256) / 2.0 - 78 if raw_rssi >= 128 else raw_rssi / 2.0 - 78
        monitor_id = int.from_bytes(packet[9:12], byteorder='little')
        cmd = packet[12]
        status_byte = packet[13]
        # Decode status byte safely
        status_fields = PacketDecoder.decode_status_byte(status_byte)
        if status_fields is None:
            print("Error: Failed to decode status byte.")
            return None
        ir_id = struct.unpack('<H', packet[14:16])[0]
        version = packet[16]
        astar_id = struct.unpack('<H', packet[17:19])[0]
        lbi = struct.unpack('<H', packet[19:21])[0]
        cmd3 = packet[21:24].hex()
        time_stamp = int.from_bytes(packet[24:28], byteorder='little')
        readable_time = datetime.datetime.utcfromtimestamp(time_stamp).strftime('%Y-%m-%d %H:%M:%S')
        ekey = packet[28]
        lf_flag = packet[29]
        Alive_status = packet[30]
        Alive_fields = PacketDecoder.decode_Alive_status(Alive_status)
        door_status = packet[31]
        temp_val = struct.unpack('<h', packet[32:34])[0]  # little-endian signed short
        temperature = round(temp_val / 100.0, 2)
        #tag_mac_id = ':'.join(f'{b:02X}' for b in packet[34:40])

        # Decode status byte
        status_fields = PacketDecoder.decode_status_byte(status_byte)

        # Prepare the decoded data
        decoded_data = {
            'Device Type': 'Tag' if device_type == 0x01 else 'Monitor',
            'Tag ID': tag_id,
            #'Tag MAC ID': tag_mac_id,
            'RSSI': f"{rssi:.2f} dBm",
            'Monitor ID': monitor_id,
            'CMD': cmd,
            'Status Byte': status_fields['Status Byte'],
            'Button 4': status_fields['Button 4'],
            'Button 3': status_fields['Button 3'],
            'Button 2': status_fields['Button 2'],
            'Button 1': status_fields['Button 1'],
            'Motion Flag': status_fields['Motion Flag'],
            'Retry Count': status_fields['Retry Count'],
            'Reserved': status_fields['Reserved'],
            'IR ID': ir_id,
            'Version': version,
            'Astar ID': astar_id,
            'LBI': lbi,
            'CMD3': cmd3,
            'Timestamp': readable_time,
            'EKEY': ekey,
            'LF Flag': lf_flag,
            'Sleep' : Alive_fields['Sleep_status'],
            'Data_index' : Alive_fields['Index'],
            'Door_status' : door_status,
            'Temperature': temperature
        }

        # Log the decoded data in a single line
        logging.info(json.dumps(decoded_data))

        # Print decoded values (optional)
        for key, value in decoded_data.items():
            print(f"{key}: {value}")

        return decoded_data


class PacketCapture:
    def __init__(self, host='192.168.1.10', port=7167, timeout=10):
        self.host = host
        self.port = port
        self.timeout = timeout

    def capture(self, timeout_type="short"):
        self.captured_packets = []
        start_time = time.time()

        # Define timeout values based on input type
        timeouts = {
            "short": 5,  # Short timeout: 5 seconds
            "medium": 15,  # Medium timeout: 10 seconds
            "long": 80, # Long timeout: 80 seconds
            "move": 0.1
        }

        # Get the timeout value based on the input argument, default to "short"
        self.timeout = timeouts.get(timeout_type, 2)

        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.bind((self.host, self.port))
            s.settimeout(self.timeout)

            while time.time() - start_time < self.timeout:
                try:
                    packet, addr = s.recvfrom(1024)  # Buffer size is 1024 bytes
                    print(f"Packet from {addr}")
                    print(f"Raw Data: {packet.hex()}")
                    self.captured_packets.append(packet)
                except socket.timeout:
                    print("Socket timed out, ending capture.")
                    break

    def validate_packet(self, decoded_packet, expected_values):
        """
        Validates if all key-value pairs in expected_values match those in decoded_packet.
        Adds a tolerance of 100 for 'LBI' and logs debug information to a file.

        Args:
            decoded_packet (dict): The actual decoded packet data.
            expected_values (dict): The expected key-value pairs to validate.

        Returns:
            bool: True if all expected values match (within tolerance for LBI), False otherwise.
        """


        mismatches = {}

        for key, expected_value in expected_values.items():
            actual_value = decoded_packet.get(key, None)

            if key == "LBI":
                # Apply tolerance of 100 for 'LBI'
                if actual_value is None or not (expected_value - 100 <= actual_value <= expected_value + 100):
                    mismatches[key] = {
                        "expected": f"{expected_value} ± 100",
                        "actual": actual_value
                    }
                    logging.debug(f"LBI mismatch: Expected {expected_value} ± 100, Got {actual_value}")
            else:
                # Exact match for other keys
                if actual_value != expected_value:
                    mismatches[key] = {
                        "expected": expected_value,
                        "actual": actual_value
                    }
                    logging.info(f"Mismatch for '{key}': Expected {expected_value}, Got {actual_value}")

        if mismatches:
            # Log all mismatches for debugging
            logging.info("Validation failed with mismatches:")
            print("Validation failed with mismatches:")
            for key, mismatch in mismatches.items():
                logging.info(f"Key: {key}, Expected: {mismatch['expected']}, Actual: {mismatch['actual']}")
                print(f"Key: {key}, Expected: {mismatch['expected']}, Actual: {mismatch['actual']}")
            return False

        logging.info("Validation successful!")
        return True

    def process_packets(self, expected_values):
        validation_result = False

        i = 0
        while i < len(self.captured_packets):
            if i + 1 < len(self.captured_packets):
                combined_packet = self.captured_packets[i] + self.captured_packets[i + 1]
                print("\nCombined Packet:")
                print(combined_packet.hex())

                if len(combined_packet) >= 13:  # Ensure at least header length
                    header = combined_packet[:13]
                    data_length,checksum = PacketDecoder.decode_header(header)
                    if data_length is None:
                        print("Skipping packet due to header decode failure.")
                        i += 2
                        continue

                    num_location_packets = data_length // 34
                    print(f"\nNumber of Location Packets: {num_location_packets}")

                    for j in range(num_location_packets):
                        start_idx = 13 + j * 34
                        end_idx = start_idx + 34
                        if end_idx > len(combined_packet):
                            print(f"Skipping incomplete location packet {j + 1}")
                            break

                        location_packet = combined_packet[start_idx:end_idx]
                        print(f"\nLocation Packet {j + 1}:")
                        decoded_packet = PacketDecoder.decode_location_packet(location_packet,checksum)

                        if decoded_packet is None:
                            print("Skipping invalid location packet.")
                            continue

                        validation_result = self.validate_packet(decoded_packet, expected_values)
                        if validation_result:
                            print("Expected parameters matched in this location packet!")
                            return True
                else:
                    print("Combined packet does not contain enough data for header.")
            i += 2

        return validation_result




class MainProcess:
    def __init__(self):
        self.serial_conn = SerialConnection(port=Arduino_serial_port, baud_rate=9600)
        self.packet_capture = PacketCapture()
        print("Initializing serial connection...")
        self.serial_conn.init_serial()

    def button_api(self, serial_message, expected_values):


        print(f"Sending command '{serial_message}' to Arduino...")
        self.serial_conn.send_command(serial_message)
        time.sleep(0.5)
        print("Capturing packets...")
        self.packet_capture.capture("short")

        print("Processing and validating captured packets...")
        is_valid = self.packet_capture.process_packets(expected_values)


        return is_valid

    def Location_api(self, serial_message, expected_values):


        print(f"Sending command '{serial_message}' to Arduino...")
        self.serial_conn.send_command(serial_message)
        time.sleep(0.5)
        print("Capturing packets...")
        self.packet_capture.capture("medium")

        print("Processing and validating captured packets...")
        is_valid = self.packet_capture.process_packets(expected_values)


        return is_valid

    def Command_api(self, serial_message, expected_values):
        print(f"Sending command '{serial_message}' to Arduino...")
        self.serial_conn.send_command(serial_message)
        time.sleep(0.5)
        print("Capturing packets...")
        self.packet_capture.capture("long")

        print("Processing and validating captured packets...")
        is_valid = self.packet_capture.process_packets(expected_values)

        return is_valid

    def get_blink_count(self, timeout=10):
        """
        Reads from the serial connection until timeout or a blink count is found.

        Args:
            self: Object containing `serial_conn` with a `.readline()` method.
            timeout (int): How long to wait (in seconds) before giving up.

        Returns:
            int | None: The blink count if found, otherwise None.
        """
        start_time = time.time()
        while time.time() - start_time < timeout:
            line = self.serial_conn.readline().decode(errors="ignore").strip()
            if not line:
                continue

            print(f"Serial Output: {line}")

            if "Blinks:" in line:
                try:
                    return int(line.split("Blinks:")[1].strip())
                except (IndexError, ValueError):
                    print("Failed to parse blink count from:", line)
                    return None
        return None




process = MainProcess()
if __name__ == "__main__":


    print("Executing G4Version_01 : To validate tag FW version  ")
    logging.info("Executing Version_01 : To validate tag FW version \n  ")
    time.sleep(3)
    expected_values = {
        'Tag ID': g4_tag1_id,
        'Version': 4
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)
    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: Version_01 Test PASS")
        Report("Version_01", 1)
        Output(f"Iteration{i}: Version_01 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: Version_01 Test FAIL")
        Report("Version_01", 0)
        Output(f"Iteration{i}: Version_01 Test FAIL")
    #-----------------------------------------------------------------------------------
    print("Executing G4_DeviceType_02 : To validate device type")
    logging.info("Executing G4_DeviceType_02 : To validate device type \n  ")
    time.sleep(3)
    expected_values = {
        'Tag ID': g4_tag1_id,
        'Device Type': 1  # 1 for Tag, 0 for Monitor                #tag type 35, SW , HW version
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)
    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: DeviceType_02 Test PASS")
        Report("DeviceType_02", 1)
        Output(f"Iteration{i}: DeviceType_02 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: DeviceType_02 Test FAIL")
        Report("DeviceType_02", 0)
        Output(f"Iteration{i}: DeviceType_02 Test FAIL")
    #--------------------------------------------------------------------------------------
    print("Executing G4_TagId_03 : To validate tag id")
    logging.info("Executing G4_TagId_03 : To validate tag id \n ")
    time.sleep(3)
    expected_values = {
        'Tag ID': 8457892
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)
    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: TagId_03 Test PASS")
        Report("TagId_02", 1)
        Output(f"Iteration{i}: TagId_03 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: TagId_03 Test FAIL")
        Report("TagId_03", 0)
        Output(f"Iteration{i}: TagId_03 Test FAIL")
    #---------------------------------------------------------------------------------
    print("Executing G4LBI_04 : To validate tag LBI value ")
    logging.info("Executing LBI_04 : To validate tag LBI value \n  ")
    time.sleep(3)
    expected_values = {
        'Tag ID': g4_tag1_id,
        'LBI': 3550 #3550±100
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)
    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: LBI_04 Test PASS")
        Report("G4LBI_04", 1)
        Output(f"Iteration{i}: LBI_04 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: LBI_04 Test FAIL")
        Report("G4LBI_04", 0)
        Output(f"Iteration{i}: LBI_04 Test FAIL")
    #--------------------------------------------------------------------------------------
    print("Executing ButtonPress_05 : To validate tag Button Press")
    logging.info("Executing ButtonPress_05 : To validate tag Button Press \n  ")
    time.sleep(3)
    expected_values = {
        'Tag ID': g4_tag1_id,
        'Button 1': '1'  # long and short press
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)
    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: ButtonPress_05 Test PASS")
        Report("ButtonPress_05", 1)
        Output(f"Iteration{i}: ButtonPress_05 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: ButtonPress_05 Test FAIL")
        Report("ButtonPress_05", 0)
        Output(f"Iteration{i}: ButtonPress_05 Test FAIL")
    #-----------------------------------------------------------------------------------
    print("Executing G4DoorStatus_06 : To validate door status")
    logging.info("Executing G4DoorStatus_06 : To validate door status\n")
    time.sleep(3)

    # Door_status: 1 = CLOSED, 2 = OPEN
    expected_values = {
        'Tag ID': g4_tag1_id,
        'Door_status': 2
    }

    result = process.button_api(serial_message="t1lp", expected_values=expected_values)   #automation for profile configuration

    i = 1
    if result:
        door_status = expected_values['Door_status']

        if door_status == 2:
            print("Door is OPEN as expected.")
            print(f"Iteration{i}: DoorStatus_06 Test PASS")
            Report("G4DoorStatus_06", 1)
            Output(f"Iteration{i}: DoorStatus_06 Test PASS")

        elif door_status == 1:
            print("Door is CLOSED as expected.")
            print(f"Iteration{i}: DoorStatus_06 Test PASS")
            Report("G4DoorStatus_06", 1)
            Output(f"Iteration{i}: DoorStatus_06 Test PASS")

        else:
            print(f"Unexpected Door_status value: {door_status}")
            print(f"Iteration{i}: DoorStatus_06 Test FAIL")
            Report("G4DoorStatus_06", 0)
            Output(f"Iteration{i}: DoorStatus_06 Test FAIL")

    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: DoorStatus_06 Test FAIL")
        Report("G4DoorStatus_06", 0)
        Output(f"Iteration{i}: DoorStatus_06 Test FAIL")
    #-----------------------------------------------------------------------------------
    print("Executing CurrentTempValue_07 : To validate temperature display")
    logging.info("Executing CurrentTempValue_07 : To validate temperature display\n")
    time.sleep(3)

    i = 1
    expected_temp = 24.8  #To Do: Need to automate temp sensor to read current room temp and getting that value here

    expected_values = {
        "Temperature": expected_temp
    }
    result = process.button_api(serial_message="t1lp", expected_values=expected_values)

    if result:
        print(f"Iteration{i}: Temperature read → {expected_temp} °C")
        logging.info(f"Iteration{i}: Temperature read: {expected_temp} °C")
        print(f"Iteration{i}: CurrentTempValue_07 Test PASS (Verify manually on display)")
        Report("CurrentTempValue_07", 1)
        Output(f"Iteration{i}: CurrentTempValue_07 Test PASS")
    else:
        print(f"Iteration{i}: Temperature mismatch or not found → FAIL")
        logging.error("Temperature mismatch or missing in packet")
        Report("CurrentTempValue_07", 0)
        Output(f"Iteration{i}: CurrentTempValue_07 Test FAIL")
    #-------------------------------------------------------------------------------------------
    #Factory Sleep

    print("Executing Move_FactorySleep_and_MonitorLED_08: To validate move tag into Factory sleep and monitor LED blinks")
    logging.info(
        "Executing Move_FactorySleep_and_MonitorLED_08: To validate move tag into Factory sleep and monitor LED blinks\n")
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Move tag into factory sleep
    result = process.button_api(serial_message="t1ho", expected_values=expected_values)
    time.sleep(0.5)

    # Immediately check LED blinks
    blink_count = process.get_blink_count(timeout=5)

    i = 1
    if result and blink_count == 3:
        print("Move_FactorySleep_and_MonitorLED_08: PASS")
        Report("Move_FactorySleep_and_MonitorLED_08", 1)
        Output(f"Iteration{i}: Move_FactorySleep_and_MonitorLED_08 Test PASS")
    elif blink_count is None:
        print("Move_FactorySleep_and_MonitorLED_08: FAIL → No blink data received")
        Report("Move_FactorySleep_and_MonitorLED_08", 0)
        Output(f"Iteration{i}: Move_FactorySleep_and_MonitorLED_08 Test FAIL (No blink data)")
    else:
        print(f"Move_FactorySleep_and_MonitorLED_08: FAIL → Blinks detected: {blink_count}")
        Report("Move_FactorySleep_and_MonitorLED_08", 0)
        Output(f"Iteration{i}: Move_FactorySleep_and_MonitorLED_08 Test FAIL (Blinks: {blink_count})")

    # --------------------------------------------------------------------------------------------------
    print("Executing ButtonPress_FactorySleep_09: To validate button press when tag is in Factory sleep")
    logging.info(
        "Executing ButtonPress_FactorySleep_09: To validate button press when tag is in Factory sleep\n"
    )
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Precondition: Tag should be in Factory sleep
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)

    # Capture packets after button press
    packets_received = process.packet_capture.captured_packets
    i = 1
    if not packets_received:
        print("No location data received after short press in factory sleep: PASS")
        print(f"Iteration{i}: ButtonPress_FactorySleep_09 Test PASS")
        Report("ButtonPress_FactorySleep_09", 1)
        Output(f"Iteration{i}: ButtonPress_FactorySleep_09 Test PASS")
    else:
        print("Location data received after short press in factory sleep: FAIL")
        print(f"Iteration{i}: ButtonPress_FactorySleep_09 Test FAIL")
        Report("ButtonPress_FactorySleep_09", 0)
        Output(f"Iteration{i}: ButtonPress_FactorySleep_09 FAIL")
    # --------------------------------------------------------------------------------------------------
    print(
        "Executing MultipleButtonPress_FactorySleep_10: To validate multiple button press when tag is in Factory sleep")
    logging.info(
        "Executing MultipleButtonPress_FactorySleep_10: To validate multiple button press when tag is in Factory sleep\n"
    )
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Precondition: Tag should be in Factory sleep
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)

    # Capture packets after button press
    packets_received = process.packet_capture.captured_packets
    i = 1
    if not packets_received:
        print("No location data received after short press in factory sleep: PASS")
        print(f"Iteration{i}: MultipleButtonPress_FactorySleep_10 Test PASS")
        Report("MultipleButtonPress_FactorySleep_10", 1)
        Output(f"Iteration{i}:MultipleButtonPress_FactorySleep_10 Test PASS")
    else:
        print("Location data received after short press in factory sleep: FAIL")
        print(f"Iteration{i}: MultipleButtonPress_FactorySleep_10 Test FAIL")
        Report("MultipleButtonPress_FactorySleep_10", 0)
        Output(f"Iteration{i}: MultipleButtonPress_FactorySleep_10 FAIL")
    # ---------------------------------------------------------------------------------------------------
    print(
        "Executing Wakeup_FactorySleep_and_MonitorLED_11: To validate wake up from Factory sleep and monitor LED blinks")
    logging.info(
        "Executing Wakeup_FactorySleep_and_MonitorLED_11: To validate wake up from Factory sleep and monitor LED blinks\n")
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Wake up from Factory sleep and validate packet
    result = process.button_api(serial_message="t1ho", expected_values=expected_values)
    time.sleep(0.5)

    # Immediately check LED blinks
    blink_count = process.get_blink_count(timeout=5)

    i = 1
    if result and blink_count == 3:
        print("Wakeup_FactorySleep_and_MonitorLED_11: PASS")
        Report("Wakeup_FactorySleep_and_MonitorLED_11", 1)
        Output(f"Iteration{i}: Wakeup_FactorySleep_and_MonitorLED_11 Test PASS")
    elif not result:
        print("Wakeup_FactorySleep_and_MonitorLED_11: FAIL → Packet validation failed")
        Report("Wakeup_FactorySleep_and_MonitorLED_11", 0)
        Output(f"Iteration{i}: Wakeup_FactorySleep_and_MonitorLED_11 Test FAIL (Packet validation failed)")
    elif blink_count is None:
        print("Wakeup_FactorySleep_and_MonitorLED_11: FAIL → No blink data received")
        Report("Wakeup_FactorySleep_and_MonitorLED_11", 0)
        Output(f"Iteration{i}: Wakeup_FactorySleep_and_MonitorLED_11 Test FAIL (No blink data)")
    else:
        print(f"Wakeup_FactorySleep_and_MonitorLED_11: FAIL → Blinks detected: {blink_count}")
        Report("Wakeup_FactorySleep_and_MonitorLED_11", 0)
        Output(f"Iteration{i}: Wakeup_FactorySleep_and_MonitorLED_11 Test FAIL (Blinks: {blink_count})")

    # ----------------------------------------------------------------------------------
    print("Executing Wakeup_FactorySleep&ButtonPress_12: To validate button press after waking up from Factory sleep")
    logging.info(
        "Executing Wakeup_FactorySleep&ButtonPress_12: To validate button press after waking up from Factory sleep\n"
    )
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Precondition: Tag should be in Factory sleep
    process.button_api(serial_message="t1sp", expected_values=expected_values)
    time.sleep(0.5)

    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: Wakeup_FactorySleep&ButtonPress_12 Test PASS")
        Report("Wakeup_FactorySleep&ButtonPress_12", 1)
        Output(f"Iteration{i}:Wakeup_FactorySleep&ButtonPress_12 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: Wakeup_FactorySleep&ButtonPress_12 Test FAIL")
        Report("Wakeup_FactorySleep&ButtonPress_12", 0)
        Output(f"Iteration{i}: Wakeup_FactorySleep&ButtonPress_12 FAIL")
    #-----------------------------------------------------------------------------------------------------
    print("Executing ButtonPress_InventorySleep_13: To validate button press when tag is in Inventory sleep")
    logging.info(
        "Executing ButtonPress_InventorySleep_13: To validate button press when tag is in Inventory sleep\n"
    )
    time.sleep(3)

    # Precondition: Tag should be in Inventory sleep
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)

    # Capture packets after button press
    packets_received = process.packet_capture.captured_packets
    i = 1
    if not packets_received:
        print("No location data received after short press in factory sleep: PASS")
        print(f"Iteration{i}: ButtonPress_InventorySleep_13 Test PASS")
        Report("ButtonPress_InventorySleep_13", 1)
        Output(f"Iteration{i}:ButtonPress_InventorySleep_13 Test PASS")
    else:
        print("Location data received after short press in factory sleep: FAIL")
        print(f"Iteration{i}: ButtonPress_InventorySleep_13 Test FAIL")
        Report("ButtonPress_InventorySleep_13", 0)
        Output(f"Iteration{i}: ButtonPress_InventorySleep_13 FAIL")
    #------------------------------------------------------------------------------------------
    print("Executing MultipleButtonPress_InventorySleep_14: To validate button press multiple times when tag is in Inventory sleep")
    logging.info(
        "Executing MultipleButtonPress_InventorySleep_14: To validate button press multiple times when tag is in Inventory sleep\n"
    )
    time.sleep(3)

    # Precondition: Tag should be in Inventory sleep
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)
    process.button_api(serial_message="t1sp", expected_values={})
    time.sleep(0.5)

    # Capture packets after button press
    packets_received = process.packet_capture.captured_packets
    i = 1
    if not packets_received:
        print("No location data received after short press in factory sleep: PASS")
        print(f"Iteration{i}: MultipleButtonPress_InventorySleep_14 Test PASS")
        Report("MultipleButtonPress_InventorySleep_14", 1)
        Output(f"Iteration{i}:MultipleButtonPress_InventorySleep_14 Test PASS")
    else:
        print("Location data received after short press in factory sleep: FAIL")
        print(f"Iteration{i}: MultipleButtonPress_InventorySleep_14 Test FAIL")
        Report("MultipleButtonPress_InventorySleep_14", 0)
        Output(f"Iteration{i}: MultipleButtonPress_InventorySleep_14 FAIL")
    #------------------------------------------------------------------------------------------
    print(
        "Executing Wakeup_InventorySleep_and_MonitorLED_15: To validate wake up from Inventory sleep and monitor LED blinks")
    logging.info(
        "Executing Wakeup_InventorySleep_and_MonitorLED_15: To validate wake up from Inventory sleep and monitor LED blinks\n")
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'version': 4
    }

    # Wake up from Inventory sleep and validate packet
    result = process.button_api(serial_message="t1ho", expected_values=expected_values)
    time.sleep(0.5)

    # Immediately check LED blinks
    blink_count = process.get_blink_count(timeout=5)

    i = 1
    if result and blink_count == 3:
        print("Wakeup_InventorySleep_and_MonitorLED_15: PASS")
        Report("Wakeup_InventorySleep_and_MonitorLED_15", 1)
        Output(f"Iteration{i}: Wakeup_InventorySleep_and_MonitorLED_15 Test PASS")
    elif not result:
        print("Wakeup_InventorySleep_and_MonitorLED_15: FAIL → Packet validation failed")
        Report("Wakeup_InventorySleep_and_MonitorLED_15", 0)
        Output(f"Iteration{i}: Wakeup_InventorySleep_and_MonitorLED_15 Test FAIL (Packet validation failed)")
    elif blink_count is None:
        print("Wakeup_InventorySleep_and_MonitorLED_15: FAIL → No blink data received")
        Report("Wakeup_InventorySleep_and_MonitorLED_15", 0)
        Output(f"Iteration{i}: Wakeup_InventorySleep_and_MonitorLED_15 Test FAIL (No blink data)")
    else:
        print(f"Wakeup_InventorySleep_and_MonitorLED_15: FAIL → Blinks detected: {blink_count}")
        Report("Wakeup_InventorySleep_and_MonitorLED_15", 0)
        Output(f"Iteration{i}: Wakeup_InventorySleep_and_MonitorLED_15 Test FAIL (Blinks: {blink_count})")

    #--------------------------------------------------------------------------------------
    print("Executing WakeupInventorySleep&ButtonPress_16: To validate button press after waking up from Inventory sleep")
    logging.info(
        "Executing WakeupInventorySleep&ButtonPress_16: To validate button press after waking up from Inventory sleep\n"
    )
    time.sleep(3)

    expected_values = {
        'Tag ID': g4_tag1_id,
        'Button 1': '1'
    }

    #button press
    process.button_api(serial_message="t1sp", expected_values=expected_values)
    time.sleep(0.5)

    i = 1
    if result:
        print("Packet validation successful.")
        print(f"Iteration{i}: Wakeup_InventorySleep_16 Test PASS")
        Report("Wakeup_InventorySleep_16", 1)
        Output(f"Iteration{i}: Wakeup_InventorySleep_16 Test PASS")
    else:
        print("Packet validation failed.")
        print(f"Iteration{i}: Wakeup_InventorySleep_16 Test FAIL")
        Report("Wakeup_InventorySleep_16", 0)
        Output(f"Iteration{i}: Wakeup_InventorySleep_16 Test FAIL")

    #----------------------------------------------------------------------------------------------


#integrate current capture for all cases
#use pywin to automate inventory sleep tool
        
        process.serial_conn.close()

